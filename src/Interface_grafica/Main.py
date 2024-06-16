from tkinter import *
from typing import Tuple
import customtkinter
import os
import sys
# Pega o diretorio pai do arquivo
diretorioPai = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#fornece o caminho
sys.path.append(diretorioPai)
from objects.planilha import iniciar_planilhas
from openpyxl import load_workbook

casa_path = os.path.abspath("Casa.xlsx")
casa_comodos_path = os.path.abspath("Casa_comodos.xlsx")

if not (os.path.exists(casa_path)) and os.path.exists(os.path.dirname(casa_comodos_path)):
    iniciar_planilhas()

from Interface_grafica.HeaderFrame import Header
from Interface_grafica.ComodoFrame import ComodoFrame
from Interface_grafica.DispositivosFrame import DispositivosFrame
from Interface_grafica.BotaoComodo import BotaoComodo
from Interface_grafica.LampadasFrame import LampadasFrame
from Interface_grafica.BotaoLampada import BotaoLampada
from Interface_grafica.BotaoDispositivos import BotaoDispositivo
from Interface_grafica.ArCondicionadoFrame import ArCondicionadoFrame
from Interface_grafica.BotaoArCondicionado import BotaoArCondicionado
from Interface_grafica.JanelaFrame import JanelaFrame
from Interface_grafica.BotaoJanela import BotaoJanela
from Interface_grafica.CortinaFrame import CortinaFrame
from Interface_grafica.BotaoCortina import BotaoCortina
from Interface_grafica.BotaoAdd import BotaoAdd
from Interface_grafica.BotaoRemove import BotaoRemove
from Interface_grafica.ComodoRemoveFrame import ComodoRemoveFrame
from Interface_grafica.ConfigurarLampada import ConfigurarLampadas
from Interface_grafica.ConfigurarCortina import ConfigurarCortina
from Interface_grafica.ConfigurarJanela import ConfigurarJanela
from Interface_grafica.ConfigurarAr import ConfigurarAr
from Interface_grafica.DispostivoAdd import DispositivoAddFrame
from Interface_grafica.ComodoAddFrame import ComodoAddFrame
from objects.Casa import Casa
from objects.Comodo import Comodo
from objects.ArCondicionado import Ar_Condicionado,criar_instancia_ar_condicionado
from objects.Cortinas import Cortina,criar_instancia_cortina
from objects.Janelas import Janela,criar_instancia_janela
from objects.Lampadas import Lampadas,criar_instancia_lampada
from objects.Dispositivo import Dispositivo
from typing import Type
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.title('Automação Residencial')
        self.geometry('390x644')
        self.configure(fg_color ="#616D7A")
        self.casa = Casa("Casa")
        self.comodos = []
        self.dispositivosFrame = {}
        self.lampadasFrame ={}
        self.lampadasBotoes = {}
        self.ArCondicionadoFrame = {}
        self.ArCondicionadoBotoes = {}
        self.janelaFrame = {}
        self.janelaBotoes = {}
        self.cortinasFrame ={}
        self.cortinasBotoes = {}
        self.botoesComodo = {}
        self.lampadasConfig = {}
        self.arConfig = {}
        self.janelaConfig = {}
        self.cortinaConfig = {}
        self.frameAtual = ""
        self.comodoFrame = False
        self.contadorLampada= {}
        self.contadorArCondicionado = {}
        self.contadorJanela = {}
        self.contadorCortina = {}
        self.botaoLampada = {}
        self.botaoAr = {}
        self.botaoCortina = {}
        self.botaoJanela = {}
        self.nomeComodos=[]
        self.nomeLampadas=[]
        self.nomeArCondicionado=[]
        self.nomeJanelas=[]
        self.nomeCortina=[]
        self.botaoAddLampada = {}
        self.botaoAddAr = {}
        self.botaoAddJanela={}
        self.botaoAddCortina = {}
        self.labelDuplicataDispositivos = customtkinter.CTkLabel(self)
        self.CarregarVetores()
        self.CriarJanelas()
        
    def CarregarVetores(self):
        self.nomeComodos=[]
        self.nomeLampadas=[]
        self.nomeArCondicionado=[]
        self.nomeJanelas=[]
        self.nomeCortina=[]
        
        wb=load_workbook("Casa.xlsx")
        ws_lampadas=wb["Lampadas"]
        ws_cortinas=wb["Cortinas"]
        ws_janelas=wb["Janelas"]
        ws_ares=wb["Ares Condicionados"]
        wb_Casa_comodo=load_workbook("Casa_comodos.xlsx")
        ws=wb_Casa_comodo["Comodos"]
        
        for i,row in enumerate(ws.iter_rows(min_row=2), start=1):
            self.nomeComodos.append((f"{str(row[0].value)}",f"{str(row[1].value)}"))

        for i,row in enumerate(ws_lampadas.iter_rows(min_row=2), start=1):
            self.nomeLampadas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","",f"{str(row[4].value)}","","",f"{str(row[7].value)}"))


        for i,row in enumerate(ws_cortinas.iter_rows(min_row=2), start=1):
            self.nomeCortina.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","",f"{str(row[4].value)}","","",""))


        for i,row in enumerate(ws_ares.iter_rows(min_row=2), start=1):
            if(row[2].value=="True"):
                self.nomeArCondicionado.append((f"{str(row[0].value)}",f"{str(row[1].value)}","True",f"{str(row[3].value)}",f"{str(row[4].value)}","","",""))
            else:
                self.nomeArCondicionado.append((f"{str(row[0].value)}",f"{str(row[1].value)}","False",f"{str(row[3].value)}",f"{str(row[4].value)}","","",""))

        for i,row in enumerate(ws_janelas.iter_rows(min_row=2), start=1):
            if(row[6].value=="True"):
                self.nomeJanelas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","True",""))
            else:
                self.nomeJanelas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","False",""))


    def CriarJanelas(self):
        self.comodoFrame = ComodoFrame(self)
        self.comodoFrame.pack(side= "top")
        self.labelDuplicata = customtkinter.CTkLabel(self.comodoFrame,
                                               text= "Já existe um comodo com esse nome, insira outro",
                                               text_color= "red")
        
        self.labelInexistente = customtkinter.CTkLabel(self.comodoFrame,
                                               text= "Esse cômodo não existe",
                                               text_color= "red")
        BotaoAdicionarComodo = BotaoAdd(self.comodoFrame, label="Adicionar novo cômodo")
        BotaoAdicionarComodo.configure(command= lambda:self.AdicionarComodoBotao())
        BotaoAdicionarComodo.pack(side = "bottom",pady= (10,10))
        
        BotaoRemoverComodo = BotaoRemove(self.comodoFrame, label="Remover cômodo")
        BotaoRemoverComodo.configure(command= lambda:self.RemoverComodoBotao())
        BotaoRemoverComodo.pack(side = "bottom",pady= (10,10))
        
        
        #Criando a pagina de cada comodo
        for nomes,numero in self.nomeComodos:
            self.contadorLampada[nomes] = 0
            self.contadorArCondicionado[nomes]  = 0
            self.contadorJanela[nomes]  = 0
            self.contadorCortina[nomes]  = 0
            
            #Criando a pagina que exibe os tipos de dispositivos
            dispositivoFrame = DispositivosFrame(master = self, nome = nomes)
            self.dispositivosFrame[nomes] = dispositivoFrame
            dispositivoFrame.header.iconeBotao.configure(command = lambda: self.VoltarFrameComodos())
            
            #Criando a pagina das lampada e configurando os botoes que irão nela
            lampadaFrame = LampadasFrame(master=self)
            lampadaFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosLampadas(n))
            self.lampadasFrame[nomes] = lampadaFrame
            
            self.botaoAddLampada[nomes] = BotaoAdd(self.lampadasFrame[nomes], label="Adicionar nova lâmpada")
            self.botaoAddLampada[nomes].configure(command = lambda n=nomes: self.AdicionarLampadaBotao(n))
            self.botaoAddLampada[nomes].pack(side = "bottom",pady= (10,10))
            
            for comodo,nomeLamp,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
                if comodo == nomes:
                    self.lampadasConfig[nomeLamp] = ConfigurarLampadas(master = self,
                                                    casas = self.casa,
                                                    comodo = nomes,
                                                    nome =nomeLamp,
                                                    intensidade=intensidade,
                                                    cor = cor
                                                    )
                    self.contadorLampada[nomes]  += 1
                    self.lampadasConfig[nomeLamp].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameLampadas(n))
                    self.lampadasConfig[nomeLamp].excluir.configure(command = lambda n1=nomeLamp,n2=nomes:self.RemoverLampada(nomeLampada=n1,nomeComodo=n2))
                    self.lampadasBotoes[nomeLamp] = BotaoLampada(self.lampadasFrame[nomes],
                                                             nomeLampada=nomeLamp,
                                                             cor=cor,
                                                             intensidade=intensidade)
                    self.lampadasBotoes[nomeLamp].configure(command = lambda n=nomeLamp:self.MudarFrameConfigLampadas(n))
                    self.lampadasBotoes[nomeLamp].pack(side="top", pady= (10,10))
            
            #Criando a pagina do ar condicionado e configurando os botoes que irão nela 
            arCondicionadoFrame = ArCondicionadoFrame(master=self)
            arCondicionadoFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosAr(n))
            self.ArCondicionadoFrame[nomes] = arCondicionadoFrame
            
            self.botaoAddAr[nomes] = BotaoAdd(self.ArCondicionadoFrame[nomes], label="Adicionar novo ar")
            self.botaoAddAr[nomes].configure(command = lambda n=nomes: self.AdicionarArBotao(n))
            self.botaoAddAr[nomes].pack(side = "bottom",pady= (10,10))
             
            for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
                if comodo == nomes:
                    self.arConfig[nome] = ConfigurarAr(master = self,
                                            casas = self.casa,
                                            comodo = nomes,
                                            nome =nome,
                                            ligado = ligado,
                                            temperatura= temperatura,
                                            intensidade=intensidade,
                                                    )
                    self.contadorArCondicionado[nomes]  += 1
                    self.arConfig[nome].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameAr(n))
                    self.arConfig[nome].excluir.configure(command = lambda n1=nome,n2=nomes:self.RemoverAr(nomeAr=n1,nomeComodo=n2))
                    
                    self.ArCondicionadoBotoes[nome] = BotaoArCondicionado(self.ArCondicionadoFrame[nomes],
                                                             nomeAr=nome,
                                                             ligado=ligado,
                                                             temperatura= temperatura,
                                                             intensidade=intensidade)
                    self.ArCondicionadoBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigAr(n))
                    self.ArCondicionadoBotoes[nome].pack(side="top", pady= (10,10))
              
              
            #Criando a pagina das janelas e configurando os botoes que irão nela        
            janelaFrame = JanelaFrame(master=self)
            janelaFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosJanela(n))
            self.janelaFrame[nomes] = janelaFrame
            
            self.botaoAddJanela[nomes] = BotaoAdd(self.janelaFrame[nomes], label="Adicionar nova janela")
            self.botaoAddJanela[nomes].configure(command = lambda n=nomes: self.AdicionarJanelaBotao(n))
            self.botaoAddJanela[nomes].pack(side = "bottom",pady= (10,10))
             
            for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeJanelas:
                if comodo == nomes:
                    self.janelaConfig[nome] = ConfigurarJanela(master = self,
                                                    casas= self.casa,
                                                    comodo = nomes,
                                                    nome =nome,
                                                    abertura=abertura,
                                                    trancado = tranca
                                                    )
                    self.janelaConfig[nome].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameJanela(n))
                    self.janelaConfig[nome].excluir.configure(command = lambda n1=nome,n2=nomes:self.RemoverJanela(nomeJanela=n1,nomeComodo=n2))
                    self.contadorJanela[nomes]  += 1
                    self.janelaBotoes[nome] = BotaoJanela(self.janelaFrame[nomes],
                                                             nomeJanela=nome,
                                                             abertura=abertura,
                                                             trancado=tranca)
                    self.janelaBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigJanela(n))
                    self.janelaBotoes[nome].pack(side="top", pady= (10,10))
             
             
            #Criando a pagina das cortinas e configurando os botoes que irão nela 
            cortinaFrame = CortinaFrame(master=self)
            cortinaFrame.header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameDispositivosCortina(n))
            self.cortinasFrame[nomes] = cortinaFrame
            
            BotaoAdicionarCortina = BotaoAdd(self.cortinasFrame[nomes], label="Adicionar nova cortina")
            BotaoAdicionarCortina.pack(side = "bottom",pady= (10,10))
                    
            for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeCortina:
                if comodo == nomes:
                    self.cortinaConfig[nome] = ConfigurarCortina(master = self,
                                                    nome =nome,
                                                    abertura=abertura,
                                                    )
                    self.cortinaConfig[nome].header.iconeBotao.configure(command = lambda n=nomes: self.VoltarFrameCortina(n))
                    self.contadorCortina[nomes]  += 1
                    self.cortinasBotoes[nome] = BotaoCortina(self.cortinasFrame[nomes],
                                                             nomeCortina=nome,
                                                             abertura=abertura)
                    self.cortinasBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigCortina(n))
                    self.cortinasBotoes[nome].pack(side="top", pady= (10,10))        
                           
            
            #Criando os botões que irão na pagina dos dispositivos    
            self.botaoLampada[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Lâmpadas",
                                  numeroDispositivos= self.contadorLampada[nomes] ,
                                  caminho="src/icons/lampada.png")
            self.botaoLampada[nomes].configure(command = lambda n=nomes:self.MudarFrameLampadas(n))
            self.botaoLampada[nomes].pack(side="top", pady= (10,10))
            
            self.botaoAr[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Ar Condicionado",
                                  numeroDispositivos= self.contadorArCondicionado[nomes] ,
                                  caminho="src/icons/arCondicionado.png")
            self.botaoAr[nomes].configure(command = lambda n=nomes:self.MudarFrameAr(n))
            self.botaoAr[nomes].pack(side="top", pady= (10,10))
        
            self.botaoJanela[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Janela",
                                  numeroDispositivos= self.contadorJanela[nomes] ,
                                  caminho="src/icons/Janela.png")
            self.botaoJanela[nomes].configure(command = lambda n=nomes:self.MudarFrameJanela(n))
            self.botaoJanela[nomes].pack(side="top", pady= (10,10))
        
            self.botaoCortina[nomes] = BotaoDispositivo(self.dispositivosFrame[nomes], nomeComodo="Cortina",
                                  numeroDispositivos= self.contadorCortina[nomes] ,
                                  caminho="src/icons/cortina.png")
            self.botaoCortina[nomes].configure(command = lambda n=nomes:self.MudarFrameCortina(n))
            self.botaoCortina[nomes].pack(side="top", pady= (10,10))
            
            #Colocando os botões da pagina do cômodo
            self.botoesComodo[nomes] = BotaoComodo(self.comodoFrame, nomeComodo=nomes, numeroDispositivos=numero)
            self.botoesComodo[nomes].configure(command = lambda n=nomes:self.MudarFrameDispositivos(n))
            self.botoesComodo[nomes].pack(side="top", pady= (10,10))

    def RecarregarBotoesComodos(self):
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes].destroy()
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes] = BotaoComodo(self.comodoFrame, nomeComodo=nomes, numeroDispositivos=numero)
            self.botoesComodo[nomes].configure(command = lambda n=nomes:self.MudarFrameDispositivos(n))
            self.botoesComodo[nomes].pack(side="top", pady= (10,10))

    def RecarregarBotoesDispositivos(self,nomeComodo):
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes].destroy()
        for nomes,numero in self.nomeComodos:
            self.botoesComodo[nomes] = BotaoComodo(self.comodoFrame, nomeComodo=nomes, numeroDispositivos=numero)
            self.botoesComodo[nomes].configure(command = lambda n=nomes:self.MudarFrameDispositivos(n))
            self.botoesComodo[nomes].pack(side="top", pady= (10,10))
        
        self.botaoLampada[nomeComodo].destroy()
        self.botaoAr[nomeComodo].destroy()
        self.botaoJanela[nomeComodo].destroy()
        self.botaoCortina[nomeComodo].destroy()
        
        
        self.botaoLampada[nomeComodo] = self.botaoLampada[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Lâmpadas",
                                  numeroDispositivos= self.contadorLampada[nomeComodo],
                                  caminho="src/icons/lampada.png")
        self.botaoLampada[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameLampadas(n))
        self.botaoLampada[nomeComodo].pack(side="top", pady= (10,10))
        
     
        self.botaoAr[nomeComodo] = self.botaoAr[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Ar Condicionado",
                                  numeroDispositivos= self.contadorArCondicionado[nomeComodo],
                                  caminho="src/icons/arCondicionado.png")
        self.botaoAr[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameAr(n))
        self.botaoAr[nomeComodo].pack(side="top", pady= (10,10))
        
        
        self.botaoJanela[nomeComodo] = self.botaoJanela[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Janela",
                                  numeroDispositivos= self.contadorJanela[nomeComodo],
                                  caminho="src/icons/Janela.png")
        self.botaoJanela[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameJanela(n))
        self.botaoJanela[nomeComodo].pack(side="top", pady= (10,10))
        
        self.botaoCortina[nomeComodo] = self.botaoCortina[nomeComodo] = BotaoDispositivo(self.dispositivosFrame[nomeComodo], nomeComodo="Cortina",
                                  numeroDispositivos= self.contadorCortina[nomeComodo],
                                  caminho="src/icons/cortina.png")
        self.botaoCortina[nomeComodo].configure(command = lambda n=nomeComodo:self.MudarFrameCortina(n))
        self.botaoCortina[nomeComodo].pack(side="top", pady= (10,10))
        
    def RecarregarBotoesLampada(self,nomeComodo):
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            if comodo == nomeComodo:
                self.lampadasBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            if comodo == nomeComodo:
                self.lampadasBotoes[nome] = BotaoLampada(self.lampadasFrame[nomeComodo],
                                                             nomeLampada=nome,
                                                             cor=cor,
                                                             intensidade=intensidade)
                self.lampadasBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigLampadas(n))
                self.lampadasBotoes[nome].pack(side="top", pady= (10,10))  
                
    def RecarregarBotoesAr(self,nomeComodo):
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
            if comodo == nomeComodo:
                self.ArCondicionadoBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
            if comodo == nomeComodo:
                self.ArCondicionadoBotoes[nome] = BotaoArCondicionado(self.ArCondicionadoFrame[nomeComodo],
                                                             nomeAr=nome,
                                                             ligado=ligado,
                                                             temperatura=temperatura,
                                                             intensidade=intensidade)
                self.ArCondicionadoBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigAr(n))
                self.ArCondicionadoBotoes[nome].pack(side="top", pady= (10,10))  
                
    def RecarregarBotoesJanela(self,nomeComodo):
        self.CarregarVetores()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeJanelas:
            if comodo == nomeComodo:
                self.janelaBotoes[nome].destroy()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeJanelas:
            if comodo == nomeComodo:
                self.janelaBotoes[nome] = BotaoJanela(self.janelaFrame[nomeComodo],
                                                             nomeJanela=nome,
                                                             trancado=tranca,
                                                             abertura=abertura)
                self.janelaBotoes[nome].configure(command = lambda n=nome:self.MudarFrameConfigJanela(n))
                self.janelaBotoes[nome].pack(side="top", pady= (10,10))  
                    
    def AdicionarLampadaBotao(self,nomeComodo):
        self.inputLampadaAdd = DispositivoAddFrame(self.lampadasFrame[nomeComodo])
        self.inputLampadaAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddLampada(n))
        self.inputLampadaAdd.pack(side="bottom")

        
        
    def SubmeterAddLampada(self,nomeComodo):
        NomeLampada = self.inputLampadaAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            if nome == NomeLampada:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.lampadasFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=1,nome=NomeLampada)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        self.lampadasConfig[NomeLampada] = ConfigurarLampadas(master = self,
                                                    casas = self.casa,
                                                    comodo = nomeComodo,
                                                    nome =NomeLampada,
                                                    intensidade=0,
                                                    cor = "Branco"
                                                    )
        self.contadorLampada[nomeComodo]  += 1
        self.lampadasConfig[NomeLampada].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameLampadas(n))
        self.lampadasConfig[NomeLampada].excluir.configure(command = lambda n1=NomeLampada,n2=nomeComodo:self.RemoverLampada(nomeLampada=n1,nomeComodo=n2))
                    
        self.lampadasBotoes[NomeLampada] = BotaoLampada(self.lampadasFrame[nomeComodo],
                                                             nomeLampada=NomeLampada,
                                                             cor="Branco",
                                                             intensidade=0)
        self.lampadasBotoes[NomeLampada].configure(command = lambda n=NomeLampada:self.MudarFrameConfigLampadas(n))
        self.lampadasBotoes[NomeLampada].pack(side="top", pady= (10,10))
        self.inputLampadaAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
    
    def RemoverLampada(self,nomeLampada,nomeComodo):
        self.casa.comodos[nomeComodo].RemoverDispositivo(1,nomeLampada)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorLampada[nomeComodo] -= 1
        self.lampadasBotoes[nomeLampada].destroy()
        self.lampadasConfig[nomeLampada].pack_forget()
        self.lampadasConfig[nomeLampada].destroy()
        self.lampadasFrame[nomeComodo].pack(side="top")
        self.frameAtual=nomeComodo
        self.CarregarVetores()
        self.RecarregarBotoesLampada(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos()
        
    def RemoverAr(self,nomeAr,nomeComodo):
        self.casa.comodos[nomeComodo].RemoverDispositivo(3,nomeAr)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorArCondicionado[nomeComodo] -= 1
        self.ArCondicionadoBotoes[nomeAr].destroy()
        self.CarregarVetores()
        self.RecarregarBotoesAr(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos() 
        self.VoltarFrameAr(nomeComodo)
        
    def AdicionarArBotao(self,nomeComodo):
        self.inputArAdd = DispositivoAddFrame(self.ArCondicionadoFrame[nomeComodo])
        self.inputArAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddAr(n))
        self.inputArAdd.pack(side="bottom")    
        
    def SubmeterAddAr(self,nomeComodo):
        NomeAr = self.inputArAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeArCondicionado:
            if nome == NomeAr:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.ArCondicionadoFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=3,nome=NomeAr)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        self.arConfig[NomeAr] = ConfigurarAr(master = self,
                                            casas = self.casa,
                                            comodo = nomeComodo,
                                            nome =NomeAr,
                                            ligado = "Desligado",
                                            temperatura= 0,
                                            intensidade=0,
                                                    )
        self.contadorArCondicionado[nomeComodo]  += 1
        self.arConfig[NomeAr].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameAr(n))
        self.arConfig[NomeAr].excluir.configure(command = lambda n1=NomeAr,n2=nomeComodo:self.RemoverAr(nomeAr=n1,nomeComodo=n2))
                    
        self.ArCondicionadoBotoes[NomeAr] = BotaoArCondicionado(self.ArCondicionadoFrame[nomeComodo],
                                                             nomeAr=NomeAr,
                                                             ligado="False",
                                                             temperatura= 20,
                                                             intensidade=0)
        self.ArCondicionadoBotoes[NomeAr].configure(command = lambda n=NomeAr:self.MudarFrameConfigAr(n))
        self.ArCondicionadoBotoes[NomeAr].pack(side="top", pady= (10,10))
        self.inputArAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
        
    def AdicionarJanelaBotao(self,nomeComodo):
        self.inputJanelaAdd = DispositivoAddFrame(self.janelaFrame[nomeComodo])
        self.inputJanelaAdd.submit.configure(command = lambda n=nomeComodo:self.SubmeterAddJanela(n))
        self.inputJanelaAdd.pack(side="bottom")

        
        
    def SubmeterAddJanela(self,nomeComodo):
        NomeJanela = self.inputJanelaAdd.input.get()
        self.labelDuplicataDispositivos.pack_forget()
        for comodo,nome,ligado,temperatura,intensidade,abertura,tranca,cor in self.nomeLampadas:
            if nome == NomeJanela:
                self.labelDuplicataDispositivos = customtkinter.CTkLabel(self.janelaFrame[nomeComodo],
                                               text= "Já existe um dispositivo com esse nome, insira outro",
                                               text_color= "red")
                self.labelDuplicataDispositivos.pack(side="top")
                return
        self.casa.comodos[nomeComodo].AdicionarDispositivo(tipo=4,nome=NomeJanela)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.CarregarVetores()
        
        self.janelaConfig[NomeJanela] = ConfigurarJanela(master = self,
                                                    casas = self.casa,
                                                    comodo = nomeComodo,
                                                    nome =NomeJanela,
                                                    abertura=0,
                                                    trancado=  "False"
                                                    )
        self.contadorJanela[nomeComodo]  += 1
        self.janelaConfig[NomeJanela].header.iconeBotao.configure(command = lambda n=nomeComodo: self.VoltarFrameJanela(n))
        self.janelaConfig[NomeJanela].excluir.configure(command = lambda n1=NomeJanela,n2=nomeComodo:self.RemoverJanela(nomeJanela=n1,nomeComodo=n2))
                    
        self.janelaBotoes[NomeJanela] = BotaoJanela(self.janelaFrame[nomeComodo],
                                                             nomeJanela=NomeJanela,
                                                             trancado="True",
                                                             abertura=0)
        self.janelaBotoes[NomeJanela].configure(command = lambda n=NomeJanela:self.MudarFrameConfigJanela(n))
        self.janelaBotoes[NomeJanela].pack(side="top", pady= (10,10))
        self.inputJanelaAdd.pack_forget()
        
        self.RecarregarBotoesDispositivos(nomeComodo)
      
    def RemoverJanela(self,nomeJanela,nomeComodo):
        self.casa.comodos[nomeComodo].RemoverDispositivo(4,nomeJanela)
        self.casa.SalvarQuantidadeDeDispositivosComodo()
        self.contadorJanela[nomeComodo] -= 1
        self.janelaBotoes[nomeJanela].destroy()
        self.CarregarVetores()
        self.RecarregarBotoesJanela(nomeComodo)
        self.RecarregarBotoesDispositivos(nomeComodo)
        self.RecarregarBotoesComodos() 
        self.VoltarFrameJanela(nomeComodo)  
        
    def AdicionarComodoBotao(self):
        self.inputComodoAdd = ComodoAddFrame(self.comodoFrame)
        self.inputComodoAdd.submit.configure(command = lambda:self.SubmeterAddComodo())
        self.inputComodoAdd.pack(side="bottom")    
        
    def SubmeterAddComodo(self):
        NomeComodo = self.inputComodoAdd.input.get()
        self.labelDuplicata.pack_forget()
        for nomes,numero in self.nomeComodos:
            if nomes == NomeComodo:
                self.labelDuplicata.pack(side = "top")
                return
        
        self.casa.AdicionarComodo(NomeComodo)
        self.CarregarVetores()
        
        
        #Criando o botao do novo comodo
        self.botoesComodo[NomeComodo] = BotaoComodo(self.comodoFrame, nomeComodo=NomeComodo, numeroDispositivos=0)
        self.botoesComodo[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameDispositivos(n))
        self.botoesComodo[NomeComodo].pack(side="top", pady= (10,10))

        #Criando o frame para os dispostivos do novo comodo        
        dispositivoFrame = DispositivosFrame(master = self, nome = NomeComodo)
        self.dispositivosFrame[NomeComodo] = dispositivoFrame
        dispositivoFrame.header.iconeBotao.configure(command = lambda: self.VoltarFrameComodos())
        
        #Criando os botões que irão na pagina dos dispositivos
        self.contadorLampada[NomeComodo] = 0
        self.contadorArCondicionado[NomeComodo]  = 0
        self.contadorJanela[NomeComodo]  = 0
        self.contadorCortina[NomeComodo]  = 0    
        self.botaoLampada[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Lâmpadas",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/lampada.png")
        self.botaoLampada[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameLampadas(n))
        self.botaoLampada[NomeComodo].pack(side="top", pady= (10,10))
        lampadaFrame = LampadasFrame(master=self)
        lampadaFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosLampadas(n))
        self.lampadasFrame[NomeComodo] = lampadaFrame
        
        self.botaoAddLampada[NomeComodo] = BotaoAdd(self.lampadasFrame[NomeComodo], label="Adicionar nova lâmpada")
        self.botaoAddLampada[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarLampadaBotao(n))
        self.botaoAddLampada[NomeComodo].pack(side = "bottom",pady= (10,10))
        
            
        self.botaoAr[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Ar Condicionado",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/arCondicionado.png")
        self.botaoAr[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameAr(n))
        self.botaoAr[NomeComodo].pack(side="top", pady= (10,10))
        arCondicionadoFrame = ArCondicionadoFrame(master=self)
        arCondicionadoFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosAr(n))
        self.ArCondicionadoFrame[NomeComodo] = arCondicionadoFrame
        
        self.botaoAddAr[NomeComodo] = BotaoAdd(self.ArCondicionadoFrame[NomeComodo], label="Adicionar novo ar")
        self.botaoAddAr[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarArBotao(n))
        self.botaoAddAr[NomeComodo].pack(side = "bottom",pady= (10,10))
        
        
        self.botaoJanela[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Janela",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/Janela.png")
        self.botaoJanela[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameJanela(n))
        self.botaoJanela[NomeComodo].pack(side="top", pady= (10,10))
        
        janelaFrame = JanelaFrame(master=self)
        janelaFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosJanela(n))
        self.janelaFrame[NomeComodo] = janelaFrame
        
        self.botaoAddJanela[NomeComodo] = BotaoAdd(self.janelaFrame[NomeComodo], label="Adicionar nova janela")
        self.botaoAddJanela[NomeComodo].configure(command = lambda n=NomeComodo: self.AdicionarJanelaBotao(n))
        self.botaoAddJanela[NomeComodo].pack(side = "bottom",pady= (10,10))
        
        
        self.botaoCortina[NomeComodo] = BotaoDispositivo(self.dispositivosFrame[NomeComodo], nomeComodo="Cortina",
                                  numeroDispositivos= 0,
                                  caminho="src/icons/cortina.png")
        self.botaoCortina[NomeComodo].configure(command = lambda n=NomeComodo:self.MudarFrameCortina(n))
        self.botaoCortina[NomeComodo].pack(side="top", pady= (10,10))  
        cortinaFrame = CortinaFrame(master=self)
        cortinaFrame.header.iconeBotao.configure(command = lambda n=NomeComodo: self.VoltarFrameDispositivosCortina(n))
        self.cortinasFrame[NomeComodo] = cortinaFrame 
        BotaoAdicionarCortina = BotaoAdd(self.cortinasFrame[NomeComodo], label="Adicionar nova cortina")
        BotaoAdicionarCortina.pack(side = "bottom",pady= (10,10))   
        
        self.inputComodoAdd.pack_forget()
        
    def RemoverComodoBotao(self):
        self.inputComodoRemove = ComodoRemoveFrame(self.comodoFrame)
        self.inputComodoRemove.submit.configure(command = lambda:self.SubmeterRemoverComodo())
        self.inputComodoRemove.pack(side="top")
        
    def SubmeterRemoverComodo(self):
        NomeComodo = self.inputComodoRemove.input.get()
        self.CarregarVetores()
        self.labelInexistente.pack_forget()
        for nomes,numero in self.nomeComodos:
            if nomes == NomeComodo:
                self.casa.RemoverComodo(NomeComodo)
                self.CarregarVetores()
                self.inputComodoRemove.pack_forget()
                self.botoesComodo[NomeComodo].pack_forget()
                self.botoesComodo[NomeComodo].destroy()
                
                #Destruindo o frame para os dispostivos do novo comodo        
                self.dispositivosFrame[NomeComodo].destroy()
                self.lampadasFrame[NomeComodo].destroy()
                self.ArCondicionadoFrame[NomeComodo].destroy()
                self.janelaFrame[NomeComodo].destroy()
                self.cortinasFrame[NomeComodo].destroy()
                return
        self.inputComodoRemove.pack_forget()
        self.labelInexistente.pack(side="top")
        self.inputComodoRemove.pack(side = "top")
        
    def MudarFrameDispositivos(self,nome):
        self.comodoFrame.pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side = "top")
        self.dispositivosFrame[nome].update()
        self.frameAtual = nome
        
    def VoltarFrameComodos(self):
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.comodoFrame.pack(side ='top')
        self.comodoFrame.update()
     
    def MudarFrameLampadas(self,nome):
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.lampadasFrame[nome].pack(side ='top')
        self.lampadasFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosLampadas(self,nome):
        self.lampadasFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def MudarFrameConfigLampadas(self,nome):
        self.lampadasFrame[self.frameAtual].pack_forget()
        self.update()
        self.lampadasConfig[nome].pack(side ='top')
        self.lampadasConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameLampadas(self,nome):
        self.lampadasConfig[self.frameAtual].pack_forget()
        self.update()
        self.lampadasFrame[nome].pack(side='top')
        self.lampadasFrame[nome].update()
        self.update()
        self.frameAtual = nome
        self.RecarregarBotoesLampada(nome)
        self.CarregarVetores()

        
    def MudarFrameAr(self,nome):
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.ArCondicionadoFrame[nome].pack(side ='top')
        self.ArCondicionadoFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosAr(self,nome):
        self.ArCondicionadoFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def MudarFrameConfigAr(self,nome):
        self.ArCondicionadoFrame[self.frameAtual].pack_forget()
        self.update()
        self.arConfig[nome].pack(side ='top')
        self.arConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameAr(self,nome):
        self.arConfig[self.frameAtual].pack_forget()
        self.update()
        self.ArCondicionadoFrame[nome].pack(side='top')
        self.ArCondicionadoFrame[nome].update()
        self.update()
        self.frameAtual = nome  
        self.RecarregarBotoesAr(nome)
        self.CarregarVetores()
         
    def MudarFrameJanela(self,nome):
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.janelaFrame[nome].pack(side ='top')
        self.janelaFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosJanela(self,nome):
        self.janelaFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.update()
        self.frameAtual = nome   
    
    def MudarFrameConfigJanela(self,nome):
        self.janelaFrame[self.frameAtual].pack_forget()
        self.update()
        self.janelaConfig[nome].pack(side ='top')
        self.janelaConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameJanela(self,nome):
        self.janelaConfig[self.frameAtual].pack_forget()
        self.update()
        self.janelaFrame[nome].pack(side='top')
        self.janelaFrame[nome].update()
        self.update()
        self.frameAtual = nome
        self.RecarregarBotoesJanela(nome)
        self.CarregarVetores()
    
    def MudarFrameCortina(self,nome):
        self.dispositivosFrame[self.frameAtual].pack_forget()
        self.update()
        self.cortinasFrame[nome].pack(side ='top')
        self.cortinasFrame[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameDispositivosCortina(self,nome):
        self.cortinasFrame[self.frameAtual].pack_forget()
        self.update()
        self.dispositivosFrame[nome].pack(side='top')
        self.dispositivosFrame[nome].update()
        self.update()
        self.frameAtual = nome     
        
    def MudarFrameConfigCortina(self,nome):
        self.cortinasFrame[self.frameAtual].pack_forget()
        self.update()
        self.cortinaConfig[nome].pack(side ='top')
        self.cortinaConfig[nome].update()
        self.frameAtual = nome
        self.update()
        
    def VoltarFrameCortina(self,nome):
        self.cortinaConfig[self.frameAtual].pack_forget()
        self.update()
        self.cortinasFrame[nome].pack(side='top')
        self.cortinasFrame[nome].update()
        self.update()
        self.frameAtual = nome      

        
app = App()
app.mainloop()