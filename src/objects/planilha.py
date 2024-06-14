# esse arquivo tem como função criar a planilha inicial 
from openpyxl import Workbook

def iniciar_planilhas():
    #Cria um arquivo excel
    wb=Workbook()
    ws=wb.active

    # Cria uma sheet chamada Comodo
    ws.title="Comodo"
    # Cria quatro sheet chamadas Lampadas, Ares Condicionados,Janelas,Cortinas 
    wb.create_sheet("Lampadas")
    wb.create_sheet("Ares Condicionados")
    wb.create_sheet("Janelas")
    wb.create_sheet("Cortinas")
    # Adiciona nas sheets dos dispositivos na primeira linha "Comodo","Nome","Ligado","Temperatura","Intensidade","Abertura","Tranca","Cor".todas separadas por colunas
    ws=wb["Lampadas"]
    ws.append(["Comodo","Nome","Ligado","Temperatura","Intensidade","Abertura","Tranca","Cor"])

    ws=wb["Ares Condicionados"]
    ws.append(["Comodo","Nome","Ligado","Temperatura","Intensidade","Abertura","Tranca","Cor"])

    ws=wb["Janelas"]
    ws.append(["Comodo","Nome","Ligado","Temperatura","Intensidade","Abertura","Tranca","Cor"])

    ws=wb["Cortinas"]
    ws.append(["Comodo","Nome","Ligado","Temperatura","Intensidade","Abertura","Tranca","Cor"])

    #Salva o nome do arquivo como Casa.xlsx
    wb.save("Casa.xlsx")

    #Cria um novo arquivo exvel
    wb_Casa_comodo=Workbook()
    ws=wb_Casa_comodo.active

    # Adiciona a sheet Comodos 
    ws.title="Comodos"

    #Adiciona na primeira linha e separadas por colunas as palavras "Comodo","Numero Dispositivos".
    ws.append(["Comodo","Numero Dispositivos"])

    #Salva esse novo arquivo como Casa_comodos.xlsx
    wb_Casa_comodo.save("Casa_comodos.xlsx")

iniciar_planilhas()