import sys
import os

# Pega o diretorio pai do arquivo
diretorioPai = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#fornece o caminho
sys.path.append(diretorioPai)

from openpyxl import Workbook,load_workbook

wb=load_workbook("Casa.xlsx")

ws_lampadas=wb["Lampadas"]
ws_cortinas=wb["Cortinas"]
ws_janelas=wb["Janelas"]
ws_ares=wb["Ares Condicionados"]


wb_Casa_comodo=load_workbook("Casa_comodos.xlsx")
ws=wb_Casa_comodo["Comodos"]



nomeComodos=[]
nomeLampadas=[]
nomeArCondicionado=[]
nomeJanelas=[]
nomeCortina=[]

for i,row in enumerate(ws.iter_rows(min_row=2), start=2):
    nomeComodos.append((f"{str(row[0].value)}",f"{str(row[1].value)}"))

for i,row in enumerate(ws_lampadas.iter_rows(min_row=2), start=2):
    nomeLampadas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","",f"{str(row[4].value)}","","",f"{str(row[7].value)}"))


for i,row in enumerate(ws_cortinas.iter_rows(min_row=2), start=2):
    nomeCortina.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","",f"{str(row[4].value)}","","",""))


for i,row in enumerate(ws_ares.iter_rows(min_row=2), start=2):
    if(row[2].value=="True"):
        nomeArCondicionado.append((f"{str(row[0].value)}",f"{str(row[1].value)}","Ligado",f"{str(row[3].value)}",f"{str(row[4].value)}","","",""))
    else:
        nomeArCondicionado.append((f"{str(row[0].value)}",f"{str(row[1].value)}","Desligado",f"{str(row[3].value)}",f"{str(row[4].value)}","","",""))

for i,row in enumerate(ws_janelas.iter_rows(min_row=2), start=2):
    if(row[6].value=="True"):
        nomeJanelas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","Trancado",""))
    else:
        nomeJanelas.append((f"{str(row[0].value)}",f"{str(row[1].value)}","","","",f"{str(row[5].value)}","Destrancado",""))


#print(f"{nomeComodos}\n")
#print(f"{nomeLampadas}\n")
#print(f"{nomeArCondicionado}\n")
#print(f"{nomeJanelas}\n")
#print(f"{nomeCortina}\n")
