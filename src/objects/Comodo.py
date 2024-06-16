import sys
import os

# Pega o diretorio pai do arquivo
diretorioPai = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#fornece o caminho
sys.path.append(diretorioPai)


from Interfaces.Interfaces_arcondicionado import InterfaceAr_Condicionado
from Interfaces.Interfaces_comodo import InterfaceComodo
from Interfaces.Interfaces_cortinas import InterfaceCortinas
from Interfaces.Interfaces_dispositivo import InterfaceDispositivo
from Interfaces.Interfaces_janelas import InterfaceJanela
from Interfaces.Interfaces_lampadas import InterfaceLampadas
from objects.ArCondicionado import Ar_Condicionado,criar_instancia_ar_condicionado
from objects.Cortinas import Cortina,criar_instancia_cortina
from objects.Janelas import Janela,criar_instancia_janela
from objects.Lampadas import Lampadas,criar_instancia_lampada
from objects.Dispositivo import Dispositivo
from typing import Type
from openpyxl import Workbook,load_workbook

wb=load_workbook("Casa.xlsx")
ws=wb.active
ws_lampadas=wb["Lampadas"]
ws_cortinas=wb["Cortinas"]
ws_janelas=wb["Janelas"]
ws_ares=wb["Ares Condicionados"]

class Comodo(InterfaceComodo):
    def __init__(self,nome:str) -> None:
        self.__nome=nome
        self.__quantidade_dispositivo=0
        self.lampadas={}
        self.arescondicionados={}
        self.cortinas={}
        self.janelas={}
        self.CarregarLampadasSalvas()
        self.CarregarCortinaSalvas()
        self.CarregarAresSalvos()
        self.CarregarJanelasSalvas()
     
    def CarregarLampadasSalvas(self)->None:
          for i,row in enumerate(ws_lampadas.iter_rows(min_row=2), start=1):
            self.lampadas[f"{str(row[1].value)}"]=criar_instancia_lampada(Lampadas,f"{str(row[1].value)}",None,0)

    def CarregarCortinaSalvas(self)->None:
          for i,row in enumerate(ws_cortinas.iter_rows(min_row=2), start=1):
            self.cortinas[f"{str(row[1].value)}"]=criar_instancia_cortina(Cortina,f"{str(row[1].value)}",0)

    def CarregarAresSalvos(self)->None:
          for i,row in enumerate(ws_ares.iter_rows(min_row=2), start=1):
            self.arescondicionados[f"{str(row[1].value)}"]=criar_instancia_ar_condicionado(Ar_Condicionado,f"{str(row[1].value)}",False,0,0)

    def CarregarJanelasSalvas(self)->None:
          for i,row in enumerate(ws_janelas.iter_rows(min_row=2), start=1):
            self.janelas[f"{str(row[1].value)}"]=criar_instancia_janela(Janela,f"{str(row[1].value)}",0,False)

    #Retorna o Nome do Comodo
    def Nome(self) -> str:
        return self.__nome
    # Seta o Nome do Comodo
    def SetNome(self, nomenovo: str) -> None:
        self.__nome=nomenovo
    
    #Checa a quantidade de dispositivos presente no quarto por meio da analise do arquivo excel chamado Casa
    def QuantidadeDispositivo(self)->int:
         valor=0
         for i, row in enumerate(ws_lampadas.iter_rows(), start=2):
               if row[0].value == self.Nome()  :
                    valor+=1
         for i, row in enumerate(ws_cortinas.iter_rows(), start=2):
               if row[0].value == self.Nome()  :
                    valor+=1
         for i, row in enumerate(ws_ares.iter_rows(), start=2):
               if row[0].value == self.Nome()  :
                    valor+=1
         for i, row in enumerate(ws_janelas.iter_rows(), start=2):
               if row[0].value == self.Nome()  :
                    valor+=1
         return valor
    
    # Cria um  dispositivo de um tipo, sendo Lampada,Cortina,Arcondicionado e  Janela, 1,2,3 e 4  respectivamente e com o nome que o usuario escolher
    def AdicionarDispositivo(self, tipo: int, nome: str) -> None:
          if(tipo==1):
               self.lampadas[nome]=criar_instancia_lampada(Lampadas,nome,cor="branca",intensidade=0)
               self.__quantidade_dispositivo+=1
               ws_lampadas.append([self.__nome,f"{self.lampadas[nome].Nome()}", None,None,self.lampadas[nome].Intensidade(),None,None,self.lampadas[nome].Cor()])
          
          elif(tipo==2):
               self.cortinas[nome]=criar_instancia_cortina(Cortina,nome,0)
               self.__quantidade_dispositivo+=1
               ws_cortinas.append([self.__nome,f"{self.cortinas[nome].Nome()}", None,None,self.cortinas[nome].Intensidade(),None,None,None])
          
          elif(tipo==3):
               self.arescondicionados[nome]=criar_instancia_ar_condicionado(Ar_Condicionado,nome,False,0,0)
               self.__quantidade_dispositivo+=1
               if(self.arescondicionados[nome].Ligado()==True):
                    ws_ares.append([self.__nome,f"{self.arescondicionados[nome].Nome()}","True",self.arescondicionados[nome].Temperatura(),self.arescondicionados[nome].Intensidade(),None,None,None])
               else:
                    ws_ares.append([self.__nome,f"{self.arescondicionados[nome].Nome()}","False",self.arescondicionados[nome].Temperatura(),self.arescondicionados[nome].Intensidade(),None,None,None])
          
          elif(tipo==4):
               self.janelas[nome]=criar_instancia_janela(Janela,nome,0,False)
               self.__quantidade_dispositivo+=1
               if(self.janelas[nome].Tranca()==True):
                    ws_janelas.append([self.__nome,f"{self.janelas[nome].Nome()}",None,None,None,self.janelas[nome].Abertura(),"True",None])
               else:
                    ws_janelas.append([self.__nome,f"{self.janelas[nome].Nome()}",None,None,None,self.janelas[nome].Abertura(),"False",None])
          wb.save("Casa.xlsx")
    
     # remove o dispositivo de um tipo,sendo Lampada,Cortina,Arcondicionado e  Janela, 1,2,3 e 4  respectivamente, que possui determinado nome      
    def RemoverDispositivo(self, tipo:int, nome: str) -> None:
          if(tipo==1):
               for i in range(ws_lampadas.max_row,1,-1) :
                    if ws_lampadas.cell(row=i,column=2).value==nome:
                         ws_lampadas.delete_rows(i, 1)
                         self.__quantidade_dispositivo-=1
               if nome in self.lampadas:
                    self.lampadas.pop(nome)
                                                
          elif(tipo==2):
               for i in range(ws_cortinas.max_row,1,-1) :
                    if ws_cortinas.cell(row=i,column=2).value==nome:
                         ws_cortinas.delete_rows(i, 1)
               if nome in self.cortinas:
                    self.cortinas.pop(nome)
                    self.__quantidade_dispositivo-=1
          elif(tipo==3):
               for i in range(ws_ares.max_row,1,-1) :
                    if ws_ares.cell(row=i,column=2).value==nome:
                         ws_ares.delete_rows(i, 1)
               if nome in self.arescondicionados:
                    self.arescondicionados.pop(nome)
                    self.__quantidade_dispositivo-=1
                                                            
          elif(tipo==4):
               for i in range(ws_janelas.max_row,1,-1) :
                    if ws_janelas.cell(row=i,column=2).value==nome:
                         ws_janelas.delete_rows(i, 1)
               if nome in self.janelas:
                    self.janelas.pop(nome)
                    self.__quantidade_dispositivo-=1            
          wb.save("Casa.xlsx")

    #Configura todas as Lampadas que possuem o nome indicado no parametro alterando todos os atributos de maneira que o usuario desejar  
    def ConfigurarLampada(self,nome:str,cor:str,intensidade:int) -> None:
         if nome in self.lampadas:
               self.lampadas[nome].SetCor(cor)
               self.lampadas[nome].SetIntensidade(intensidade)
         for i,row in enumerate(ws_lampadas.iter_rows(), start=2):
               if row[0].value == self.Nome() and row[1].value == nome:
                    row[4].value = intensidade
                    row[7].value = cor
         wb.save("Casa.xlsx")

    #Configura todas as cortinas que possuem o nome indicado no parametro alterando todos os atributos de maneira que o usuario desejar   
    def ConfiguraCortina(self,nome:str,intensidade:int)->None:
         if nome in self.cortinas:
               self.cortinas[nome].SetIntensidade(intensidade)
         for i,row in enumerate(ws_cortinas.iter_rows(), start=2):
               if row[0].value == self.Nome() and row[1].value == nome:
                     row[4].value=intensidade
         wb.save("Casa.xlsx")

    #Configura todos os ares condicionados que possuem o nome indicado no parametro alterando todos os atributos de maneira que o usuario desejar       
    def ConfigurarArCondicionado(self,nome:str,ligado:str,temperatura:int,intensidade:int)->None:
          if nome in self.arescondicionados:
                self.arescondicionados[nome].SetLigado(ligado)
                self.arescondicionados[nome].SetTemperatura(temperatura)
                self.arescondicionados[nome].SetIntensidade(intensidade)
          for i,row in enumerate (ws_ares.iter_rows(),start=2):
               if row[0].value == self.Nome() and row[1].value == nome:
                     if ligado=="False": # Se o ar esta desligado não é possível a temperatura e a intensidade é 0 
                           row[2].value="False"
                           row[3].value=0
                           row[4].value=0
                     else:
                           row[2].value="True"
                           row[3].value=temperatura
                           row[4].value=intensidade                          
          wb.save("Casa.xlsx")

    #Configura todas as janelas que possuem o nome indicado no parametro alterando todos os atributos de maneira que o usuario desejar  
    def ConfigurarJanela(self,nome:str,abertura:int,tranca:bool)->None:
          if nome in self.janelas:
                self.janelas[nome].SetTranca(tranca)
                self.janelas[nome].SetAbertura(abertura)
          for i,row in enumerate (ws_janelas.iter_rows(),start=2):
               if row[0].value == self.Nome() and row[1].value == nome:
                     if tranca==False:# Se a janela está trancada não é possivel abrir a janela 
                           row[6].value="False"
                           row[5].value=0
                     else:
                           row[6].value="True"
                           row[5].value=abertura              
          wb.save("Casa.xlsx")
     
    #Apaga Todos os dispositivos referente ao comodo na planilha excel chamado Casa
    def ApagarTodosdispositivoscomodo(self)->None: 
          for i in range (ws_lampadas.max_row,1,-1):
               if ws_lampadas.cell(row=i,column=1).value==self.Nome():
                    ws_lampadas.delete_rows(i, 1)
                    
          for i in range (ws_cortinas.max_row,1,-1):
               if ws_cortinas.cell(row=i,column=1).value==self.Nome():
                    ws_cortinas.delete_rows(i, 1)
                   
          for i in range (ws_ares.max_row,1,-1):
               if ws_ares.cell(row=i,column=1).value==self.Nome():
                    ws_ares.delete_rows(i, 1)
                     
          for i in range (ws_janelas.max_row,1,-1):
               if ws_janelas.cell(row=i,column=1).value==self.Nome():
                    ws_janelas.delete_rows(i, 1)
                    
          wb.save("Casa.xlsx")

#Cria um Comodo ja com o atributo nome e com o map dos dispositivos vazios        
def criar_comodo(classe:Type[Comodo],nome:str)->Comodo:
     instancia=classe(nome)
     return instancia
