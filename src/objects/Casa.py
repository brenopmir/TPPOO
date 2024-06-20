import sys
import os

# Pega o diretório pai do arquivo
diretorioPai = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# fornece o caminho
sys.path.append(diretorioPai)

from openpyxl import Workbook,load_workbook
from Interfaces.Interfaces_casa import InterfaceCasa
from objects.Comodo import Comodo, criar_comodo

wb_Casa_comodo=load_workbook("Casa_comodos.xlsx")
ws=wb_Casa_comodo["Comodos"]

class Casa(InterfaceCasa):
    def __init__(self, nome: str) -> None:
        self.__nome = nome
        self.comodos={}
        self.CarregarComodosSalvos()
        
    #Faz o carregamento dos valores que estão salvos previamente nas planilha Comodos do arquivo Casa_comodos e salvam esses valores no dictionary dos comodos
    def CarregarComodosSalvos(self)->None:
        for i,row in enumerate(ws.iter_rows(min_row=2), start=1):
            self.comodos[f"{str(row[0].value)}"]=criar_comodo(Comodo,f"{str(row[0].value)}")
            
    #Retorna a propriedade nome da casa       
    def Nome(self) -> str:
        return self.__nome
    
    #Seta o Nome da Casa 
    def SetNome(self, nomenovo: str) -> None:
        self.__nome = nomenovo

    #Adiciona um Comodo sem nenhum dispositivo nele 
    def AdicionarComodo(self, nomedocomodo: str) -> None:
        if not self.VerificarDuplicado(nomedocomodo):
            ws.append([nomedocomodo,0])
            wb_Casa_comodo.save("Casa_comodos.xlsx")
        self.comodos[nomedocomodo] = criar_comodo(Comodo, nomedocomodo)
            
    #Remove o Comodo e todos os dispositivos contidos nele 
    def RemoverComodo(self, nomecomodo: str) -> None:
        self.comodos[nomecomodo].ApagarTodosdispositivoscomodo()
        for i in range(ws.max_row,1,-1) :
                if ws.cell(row=i,column=1).value==nomecomodo:
                    ws.delete_rows(i, 1)
        if nomecomodo in self.comodos:
                self.comodos.pop(nomecomodo)
        wb_Casa_comodo.save("Casa_comodos.xlsx") # Salva o workbook após remover o cômodo
        
        
    #Verifica se há comodos duplicados
    def VerificarDuplicado(self, nomedocomodo: str) -> bool:
        for row in ws.iter_rows(values_only=True):
            if nomedocomodo in row:
                return True
        return False

    #Salva a quantidade de dispositivos presente nos Comodos
    def SalvarQuantidadeDeDispositivosComodo(self) -> None:
      for row in range(2,(ws.max_row+1)):
        ws['B'+str(row)]=self.comodos[ws['A' + str(row)].value].QuantidadeDispositivo()
        wb_Casa_comodo.save("Casa_comodos.xlsx")